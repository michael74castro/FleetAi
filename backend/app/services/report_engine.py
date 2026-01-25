"""
FleetAI - Report Engine Service
Executes reports and handles export functionality
"""

from typing import Any, Dict, List, Optional
import logging
import os
from datetime import datetime
from pathlib import Path

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

from app.core.config import settings

logger = logging.getLogger(__name__)


class ReportEngine:
    """Engine for executing and exporting reports"""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def execute_report(
        self,
        report_config: Dict[str, Any],
        dataset_name: Optional[str],
        parameters: Optional[Dict[str, Any]] = None,
        customer_ids: Optional[List[str]] = None,
        page: int = 1,
        page_size: int = 50
    ) -> Dict[str, Any]:
        """
        Execute a report and return paginated results.

        Args:
            report_config: Report configuration
            dataset_name: Source dataset name
            parameters: Runtime parameters
            customer_ids: Customer IDs for RLS
            page: Page number
            page_size: Items per page

        Returns:
            Dict with columns, data, total_rows, and aggregations
        """
        # Get dataset info
        from app.models.report import Dataset
        from sqlalchemy import select

        result = await self.db.execute(
            select(Dataset).where(Dataset.name == dataset_name)
        )
        dataset = result.scalar_one_or_none()

        if not dataset:
            raise ValueError(f"Dataset not found: {dataset_name}")

        # Build query
        columns = report_config.get("columns", [])
        filters = report_config.get("filters", [])
        sorting = report_config.get("sorting", [])
        grouping = report_config.get("grouping", [])
        aggregations = report_config.get("aggregations", [])

        # Column list
        if columns:
            col_list = []
            for col in columns:
                field = col.get("field") if isinstance(col, dict) else col
                label = col.get("label", field) if isinstance(col, dict) else field
                col_list.append(f"[{field}]")
            select_clause = ", ".join(col_list)
        else:
            select_clause = "*"

        # Base query
        base_query = f"SELECT {select_clause} FROM {dataset.source_object}"

        # WHERE clause
        conditions = []

        # Apply config filters
        for f in filters:
            condition = self._build_filter_condition(f, parameters)
            if condition:
                conditions.append(condition)

        # Apply RLS
        if customer_ids is not None and dataset.rbac_column:
            if customer_ids:
                ids_str = ", ".join([f"'{cid}'" for cid in customer_ids])
                conditions.append(f"{dataset.rbac_column} IN ({ids_str})")
            else:
                conditions.append("1 = 0")

        if conditions:
            base_query += " WHERE " + " AND ".join(conditions)

        # GROUP BY
        if grouping:
            base_query += f" GROUP BY {', '.join(grouping)}"

        # ORDER BY
        if sorting:
            order_parts = [f"[{s['field']}] {s.get('direction', 'asc')}" for s in sorting]
            base_query += f" ORDER BY {', '.join(order_parts)}"
        else:
            base_query += " ORDER BY (SELECT NULL)"  # Required for OFFSET

        # Get total count
        count_query = f"SELECT COUNT(*) FROM ({base_query.replace('ORDER BY (SELECT NULL)', '')}) AS cnt"
        count_result = await self.db.execute(text(count_query))
        total_rows = count_result.scalar()

        # Apply pagination
        offset = (page - 1) * page_size
        paginated_query = f"""
            {base_query}
            OFFSET {offset} ROWS
            FETCH NEXT {page_size} ROWS ONLY
        """

        # Execute
        result = await self.db.execute(text(paginated_query))
        rows = result.fetchall()
        result_columns = list(result.keys())

        # Format data
        data = [dict(zip(result_columns, row)) for row in rows]

        # Calculate aggregations
        agg_results = None
        if aggregations:
            agg_results = await self._calculate_aggregations(
                dataset.source_object,
                aggregations,
                conditions
            )

        # Format column definitions
        column_defs = []
        for col in result_columns:
            col_config = next(
                (c for c in columns if (c.get("field") if isinstance(c, dict) else c) == col),
                {}
            )
            column_defs.append({
                "field": col,
                "label": col_config.get("label", col) if isinstance(col_config, dict) else col,
                "format": col_config.get("format") if isinstance(col_config, dict) else None,
                "sortable": col_config.get("sortable", True) if isinstance(col_config, dict) else True
            })

        return {
            "columns": column_defs,
            "data": data,
            "total_rows": total_rows,
            "aggregations": agg_results
        }

    async def export_report(
        self,
        execution_id: int,
        report_config: Dict[str, Any],
        dataset_name: str,
        parameters: Optional[Dict[str, Any]],
        export_format: str,
        customer_ids: Optional[List[str]]
    ) -> None:
        """
        Export report to file (runs in background).

        Args:
            execution_id: Report execution ID to update
            report_config: Report configuration
            dataset_name: Source dataset
            parameters: Runtime parameters
            export_format: Export format (excel, pdf, csv)
            customer_ids: Customer IDs for RLS
        """
        from app.models.report import ReportExecution
        from sqlalchemy import select

        try:
            # Get all data (no pagination)
            result = await self.execute_report(
                report_config=report_config,
                dataset_name=dataset_name,
                parameters=parameters,
                customer_ids=customer_ids,
                page=1,
                page_size=100000  # Max export rows
            )

            # Generate file
            export_dir = Path(settings.EXPORT_DIR)
            export_dir.mkdir(parents=True, exist_ok=True)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"report_{execution_id}_{timestamp}.{export_format}"
            filepath = export_dir / filename

            if export_format == "excel":
                await self._export_to_excel(filepath, result)
            elif export_format == "csv":
                await self._export_to_csv(filepath, result)
            elif export_format == "pdf":
                await self._export_to_pdf(filepath, result)
            else:
                raise ValueError(f"Unsupported export format: {export_format}")

            # Update execution record
            exec_result = await self.db.execute(
                select(ReportExecution).where(ReportExecution.execution_id == execution_id)
            )
            execution = exec_result.scalar_one()

            execution.status = "success"
            execution.file_path = str(filepath)
            execution.row_count = result["total_rows"]
            execution.completed_at = datetime.utcnow()

            await self.db.commit()

            logger.info(f"Report exported: {filepath}")

        except Exception as e:
            logger.error(f"Export failed for execution {execution_id}: {e}")

            exec_result = await self.db.execute(
                select(ReportExecution).where(ReportExecution.execution_id == execution_id)
            )
            execution = exec_result.scalar_one()
            execution.status = "failed"
            execution.error_message = str(e)
            execution.completed_at = datetime.utcnow()

            await self.db.commit()

    async def _export_to_excel(self, filepath: Path, data: Dict[str, Any]) -> None:
        """Export to Excel using openpyxl"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")

        columns = data["columns"]
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col.get("label", col.get("field")))
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row_data in enumerate(data["data"], 2):
            for col_idx, col in enumerate(columns, 1):
                field = col.get("field")
                value = row_data.get(field)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit columns
        for col_idx, col in enumerate(columns, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 15

        wb.save(filepath)

    async def _export_to_csv(self, filepath: Path, data: Dict[str, Any]) -> None:
        """Export to CSV"""
        import csv

        columns = data["columns"]
        fields = [col.get("field") for col in columns]
        headers = [col.get("label", col.get("field")) for col in columns]

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)

            for row_data in data["data"]:
                row = [row_data.get(field) for field in fields]
                writer.writerow(row)

    async def _export_to_pdf(self, filepath: Path, data: Dict[str, Any]) -> None:
        """Export to PDF (placeholder - requires reportlab or weasyprint)"""
        # For now, create a simple text-based PDF
        # In production, use reportlab or weasyprint for proper PDF generation

        content = "FleetAI Report\n"
        content += "=" * 50 + "\n\n"

        columns = data["columns"]
        headers = [col.get("label", col.get("field")) for col in columns]
        content += " | ".join(headers) + "\n"
        content += "-" * 50 + "\n"

        for row_data in data["data"][:100]:  # Limit for text PDF
            row = [str(row_data.get(col.get("field"), "")) for col in columns]
            content += " | ".join(row) + "\n"

        if data["total_rows"] > 100:
            content += f"\n... and {data['total_rows'] - 100} more rows\n"

        # Write as text file with .pdf extension (placeholder)
        # In production, use proper PDF library
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(content)

    def _build_filter_condition(
        self,
        filter_config: Dict[str, Any],
        parameters: Optional[Dict[str, Any]]
    ) -> Optional[str]:
        """Build SQL filter condition from config"""
        field = filter_config.get("field")
        op = filter_config.get("op", "=")
        value = filter_config.get("value")
        param = filter_config.get("param")

        # Get value from parameters if specified
        if param and parameters:
            value = parameters.get(param.replace("@", ""))

        if not field or value is None:
            return None

        op_upper = op.upper()

        if op_upper == "IN":
            if isinstance(value, list):
                values_str = ", ".join([f"'{v}'" for v in value])
                return f"[{field}] IN ({values_str})"
            return None
        elif op_upper == "BETWEEN":
            if isinstance(value, list) and len(value) == 2:
                return f"[{field}] BETWEEN '{value[0]}' AND '{value[1]}'"
            return None
        elif op_upper == "LIKE":
            return f"[{field}] LIKE '%{value}%'"
        elif op_upper == "ILIKE":
            return f"LOWER([{field}]) LIKE LOWER('%{value}%')"
        else:
            return f"[{field}] {op} '{value}'"

    async def _calculate_aggregations(
        self,
        source_object: str,
        aggregations: List[Dict[str, Any]],
        conditions: List[str]
    ) -> Dict[str, Any]:
        """Calculate aggregation values"""
        if not aggregations:
            return {}

        agg_parts = []
        for agg in aggregations:
            field = agg.get("field")
            func = agg.get("function", "SUM")
            label = agg.get("label", f"{func}_{field}")
            agg_parts.append(f"{func}([{field}]) AS [{label}]")

        query = f"SELECT {', '.join(agg_parts)} FROM {source_object}"

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        result = await self.db.execute(text(query))
        row = result.fetchone()

        if row:
            columns = list(result.keys())
            return dict(zip(columns, row))

        return {}
